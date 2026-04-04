"""Export service — Excel, CSV, JSON export from processed results."""
from __future__ import annotations
import json
from pathlib import Path
from typing import Optional
import pandas as pd
from loguru import logger

from app.models.schemas import PriceDocument, PriceItem

RESULTS_DIR = Path(__file__).resolve().parent.parent.parent / "results"


def load_result(job_id: str) -> Optional[PriceDocument]:
    result_file = RESULTS_DIR / f"{job_id}.json"
    if not result_file.exists():
        return None
    with open(str(result_file), "r", encoding="utf-8") as f:
        data = json.load(f)
    return PriceDocument(**data)


def export_excel(job_id: str) -> Optional[Path]:
    doc = load_result(job_id)
    if not doc or not doc.items:
        return None

    output_path = RESULTS_DIR / f"{job_id}_export.xlsx"

    # Detect optional columns once at document level (mirrors frontend logic)
    has_dvt2         = any(item.dvt_2 or item.don_gia_2 is not None for item in doc.items)
    has_vat_pct      = any(item.vat_pct is not None for item in doc.items)
    has_qui_cach     = any(item.qui_cach for item in doc.items)
    has_don_gia_co_vat = any(item.don_gia_co_vat is not None for item in doc.items)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:

        # ── Sheet 1: Items ─────────────────────────────────────
        rows = []
        for i, item in enumerate(doc.items, 1):
            row = {
                "STT": item.stt or i,
                "Nhóm SP": item.nhom_sp or "",
                "Mã SP": item.ma_sp or "",
                "Tên sản phẩm": item.ten_sp,
            }
            if has_qui_cach:
                row["Quy cách"] = item.qui_cach or ""
            row.update({
                "ĐVT": item.dvt or "",
                "Đơn giá": item.don_gia,
            })
            if has_don_gia_co_vat:
                row["Đơn giá có VAT"] = item.don_gia_co_vat
            if has_dvt2:
                row["ĐVT 2"] = item.dvt_2 or ""
                row["Đơn giá 2"] = item.don_gia_2
            if has_vat_pct:
                row["VAT %"] = item.vat_pct
            row["Ghi chú"] = item.ghi_chu or ""
            row["Độ tin cậy AI"] = f"{item.confidence:.0%}"
            rows.append(row)

        df_items = pd.DataFrame(rows)
        df_items.to_excel(writer, sheet_name="Bảng Giá", index=False)

        # Format worksheet
        ws = writer.sheets["Bảng Giá"]
        _format_excel_sheet(ws, df_items)

        # ── Sheet 2: Summary ───────────────────────────────────
        summary_info = [
            "Nhà cung cấp", "Ngày hiệu lực", "Giá đã có VAT",
            "Đơn vị tiền", "Số sản phẩm",
        ]
        summary_values = [
            doc.nha_cung_cap or "",
            doc.ngay_hieu_luc or "",
            "Có" if doc.gia_da_bao_gom_vat else "Không",
            doc.don_vi_tien,
            len(doc.items),
        ]

        summary_data = {"Thông tin": summary_info, "Giá trị": summary_values}
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Tóm Tắt", index=False)

    logger.info(f"Excel exported: {output_path.name}")
    return output_path


def export_csv(job_id: str) -> Optional[Path]:
    doc = load_result(job_id)
    if not doc or not doc.items:
        return None

    output_path = RESULTS_DIR / f"{job_id}_export.csv"
    rows = []
    for i, item in enumerate(doc.items, 1):
        row = {
            "stt": item.stt or i,
            "nhom_sp": item.nhom_sp or "",
            "ma_sp": item.ma_sp or "",
            "ten_sp": item.ten_sp,
            "qui_cach": item.qui_cach or "",
            "dvt": item.dvt or "",
            "don_gia": item.don_gia,
            "don_gia_co_vat": item.don_gia_co_vat,
            "dvt_2": item.dvt_2 or "",
            "don_gia_2": item.don_gia_2,
            "vat_pct": item.vat_pct,
            "ghi_chu": item.ghi_chu or "",
            "confidence": item.confidence,
        }
        rows.append(row)
    pd.DataFrame(rows).to_csv(str(output_path), index=False, encoding="utf-8-sig")
    return output_path


def _fmt_money(val: Optional[float], currency: str = "VND") -> str:
    if val is None:
        return ""
    if currency == "VND":
        return f"{val:,.0f} đ"
    return f"{val:,.2f} {currency}"


def _format_excel_sheet(ws, df: pd.DataFrame):
    """Apply basic Excel formatting."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Header row styling
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-width columns
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() or 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Zebra stripes
        light_fill = PatternFill(start_color="EBF3FA", end_color="EBF3FA", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = light_fill
    except Exception as e:
        logger.warning(f"Excel formatting failed: {e}")
